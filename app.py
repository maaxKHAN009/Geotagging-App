# app.py
from flask import Flask, render_template, request, jsonify, redirect, url_for
import folium
import os
import json
from flask_cors import CORS
import requests
import pandas as pd
from flask import send_from_directory
import threading
from datetime import datetime
from werkzeug.utils import secure_filename

app = Flask(__name__)
CORS(app)

# In-memory storage for reports (for simplicity, data resets on server restart)
REPORTS_FILE = os.path.join(app.root_path, 'reports.json')
EXCEL_FILE = os.path.join(app.root_path, 'reports.xlsx')
IMAGE_ROOT = os.path.join(app.root_path, 'report_images')
REPORT_TYPES = ['pollution', 'deforestation', 'improvement', 'other']

# Thread lock for file access
file_lock = threading.Lock()

def load_reports():
    with file_lock:
        if os.path.exists(REPORTS_FILE):
            with open(REPORTS_FILE, 'r', encoding='utf-8') as f:
                try:
                    return json.load(f)
                except Exception as e:
                    print(f"Error loading reports: {e}")
                    return []
        return []

def save_reports(reports):
    with file_lock:
        with open(REPORTS_FILE, 'w', encoding='utf-8') as f:
            json.dump(reports, f, ensure_ascii=False, indent=2)
        # Save to Excel with separate sheets and formatting
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        # Group reports by type
        categories = ['pollution', 'deforestation', 'improvement', 'other']
        grouped = {cat: [] for cat in categories}
        for r in reports:
            cat = r.get('type', 'other').lower()
            if cat not in grouped:
                grouped['other'].append(r)
            else:
                grouped[cat].append(r)

        wb = Workbook()
        wb.remove(wb.active)

        for cat in categories:
            data = grouped[cat]
            if not data:
                continue
            ws = wb.create_sheet(title=cat.capitalize())
            # Capitalize column headers
            columns = [k.capitalize() for k in data[0].keys()]
            ws.append(columns)
            for row in data:
                # Convert images list to comma-separated string for Excel
                row_data = []
                for k in row.keys():
                    if k == "images" and isinstance(row[k], list):
                        row_data.append(", ".join(row[k]))
                    else:
                        row_data.append(row.get(k, ""))
                ws.append(row_data)
            # Create table
            end_col = get_column_letter(len(columns))
            end_row = len(data) + 1
            table = Table(displayName=f"{cat.capitalize()}Table", ref=f"A1:{end_col}{end_row}")
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)
            # Autofit columns, but limit max width
            max_width = 40
            for col_idx, col in enumerate(ws.columns, 1):
                max_len = max((len(str(cell.value)) for cell in col), default=10)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)

        wb.save(EXCEL_FILE)

reports_data = load_reports()

def english_to_urdu(text_input):
    """
    Uses MyMemory API to translate English to Urdu.
    """
    try:
        url = "https://api.mymemory.translated.net/get"
        params = {"q": text_input, "langpair": "en|ur"}
        response = requests.get(url, params=params, timeout=5)
        data = response.json()
        return data['responseData']['translatedText']
    except Exception:
        return "Translation service error."

@app.route('/')
def index():
    """
    Renders the main web page and generates the Folium map.
    """
    # Approximate coordinates for Gilgit-Baltistan (Gilgit city)
    gb_coords = (35.9208, 74.3088)
    m = folium.Map(location=gb_coords, zoom_start=9, tiles="OpenStreetMap")

    # Add existing reports as markers on the map
    for report in reports_data:
        # For simplicity, using dummy lat/lon for in-memory reports
        # In a real app, you'd store actual lat/lon from user input
        # These are just to make markers appear somewhat spread out on the map
        lat = float(report['coord_x']) # Using coord_x as lat for marker
        lon = float(report['coord_y']) # Using coord_y as lon for marker
        
        popup_html = f"<b>Type:</b> {report['type']}<br><b>Description:</b> {report['description']}"
        
        color = 'blue'
        if report['type'] == 'pollution':
            color = 'red'
        elif report['type'] == 'deforestation':
            color = 'darkred'
        elif report['type'] == 'improvement':
            color = 'green'
        elif report['type'] == 'other':
            color = 'orange'

        folium.Marker(
            location=[lat, lon],
            popup=folium.Popup(popup_html, max_width=300),
            icon=folium.Icon(color=color, icon='info-sign')
        ).add_to(m)

    # Save map to a static HTML file that the iframe will load
    # Ensure 'static' directory exists in your Flask app root
    map_html_path = os.path.join(app.root_path, 'static', 'map.html')
    m.save(map_html_path)
    
    return render_template('index.html')

@app.route('/health', methods=['GET'])
def health():
    """Health check endpoint."""
    return jsonify({'status': 'ok'})

def ensure_image_folders():
    os.makedirs(IMAGE_ROOT, exist_ok=True)
    for t in REPORT_TYPES:
        os.makedirs(os.path.join(IMAGE_ROOT, t), exist_ok=True)

@app.route('/submit_report', methods=['POST'])
def submit_report():
    """
    Handles submission of new environmental reports and keeps a history in the file.
    Handles image uploads directly with the report.
    """
    try:
        report_type = request.form.get('type') or request.form.get('reportType')
        location = request.form.get('reportLocation')
        description = request.form.get('description') or request.form.get('reportDescription')
        coord_x = request.form.get('coordX')
        coord_y = request.form.get('coordY')
        # Input validation
        if not (report_type and location and description):
            print("Missing required fields:", report_type, location, description)
            return jsonify({"status": "error", "message": "Missing required fields"}), 400
        if len(description) > 1000:
            print("Description too long.")
            return jsonify({"status": "error", "message": "Description too long."}), 400
        try:
            if coord_x and coord_y:
                coord_x = float(coord_x)
                coord_y = float(coord_y)
            else:
                lat_str, lon_str = location.replace('Lat: ', '').replace('Lon: ', '').split(', ')
                coord_x = float(lat_str)
                coord_y = float(lon_str)
        except Exception as e:
            print("Invalid location format:", e)
            return jsonify({"status": "error", "message": f"Invalid location format: {str(e)}"}), 400

        # Ensure image folders exist before saving
        ensure_image_folders()
        # Debug: print uploaded files
        print("FILES RECEIVED:", request.files)
        image_files = request.files.getlist('images')
        print("FILES IN 'images':", image_files)
        if not image_files or not any(img and img.filename for img in image_files):
            print("No images received or filenames missing.")
            return jsonify({"status": "error", "message": "At least one image is required."}), 400
        image_filenames = []
        folder_name = report_type.lower() if report_type and report_type.lower() in REPORT_TYPES else 'other'
        folder_path = os.path.join(IMAGE_ROOT, folder_name)
        os.makedirs(folder_path, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        for idx, img in enumerate(image_files):
            if img and img.filename:
                ext = os.path.splitext(img.filename)[1]
                safe_name = secure_filename(f"{timestamp}_{idx+1}{ext}")
                save_path = os.path.join(folder_path, safe_name)
                try:
                    print(f"Saving image to: {save_path}")  # Debug print
                    img.save(save_path)
                    image_filenames.append(safe_name)
                except Exception as file_save_exc:
                    print(f"Failed to save image {safe_name}: {file_save_exc}")

        if not image_filenames:
            print("No images saved.")
            return jsonify({"status": "error", "message": "Failed to save images."}), 500

        new_report = {
            "type": report_type,
            "location": location,
            "description": description,
            "coord_x": coord_x,
            "coord_y": coord_y,
            "datetime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "images": image_filenames
        }
        reports = load_reports()
        reports.append(new_report)
        save_reports(reports)
        print(f"Report added: {new_report}")
        return jsonify({"status": "success", "report": new_report}), 201
    except Exception as e:
        print(f"Error in submit_report: {e}")
        return jsonify({"status": "error", "message": f"Server error: {str(e)}"}), 500

@app.route('/uploads/<report_type>/<filename>')
def uploaded_file(report_type, filename):
    folder_path = os.path.join(IMAGE_ROOT, report_type)
    return send_from_directory(folder_path, filename)

@app.route('/debug_upload', methods=['GET', 'POST'])
def debug_upload():
    if request.method == 'POST':
        files = request.files.getlist('images')
        print("DEBUG UPLOAD FILES:", files)
        for img in files:
            print("DEBUG UPLOAD FILENAME:", img.filename)
        return "Check server logs for debug info."
    return '''
    <form method="POST" enctype="multipart/form-data">
        <input type="file" name="images" multiple>
        <button type="submit">Upload</button>
    </form>
    '''

@app.route('/get_reports', methods=['GET'])
def get_reports():
    """
    Returns all submitted reports as JSON.
    """
    try:
        return jsonify(load_reports())
    except Exception as e:
        print(f"Error in get_reports: {e}")
        return jsonify({"status": "error", "message": "Could not load reports."}), 500

@app.route('/recent_reports', methods=['GET'])
def recent_reports():
    return jsonify([])

@app.route('/translate', methods=['POST'])
def translate():
    """
    Handles translation requests using MyMemory API for English to Urdu.
    """
    data = request.get_json()
    text_input = data.get('text')
    translated_text = english_to_urdu(text_input)
    return jsonify({"translated_text": translated_text})

@app.route('/export_reports_excel', methods=['GET'])
def export_reports_excel():
    """
    Exports all reports as an Excel file for download.
    """
    try:
        reports = load_reports()
        if not reports:
            return jsonify({'status': 'error', 'message': 'No reports to export.'}), 404
        df = pd.DataFrame(reports)
        excel_path = os.path.join(app.root_path, 'static', 'reports_export.xlsx')
        df.to_excel(excel_path, index=False)
        return send_from_directory(directory=os.path.join(app.root_path, 'static'), filename='reports_export.xlsx', as_attachment=True)
    except Exception as e:
        print(f"Error in export_reports_excel: {e}")
        return jsonify({'status': 'error', 'message': 'Failed to export reports.'}), 500

@app.route('/logo.png')
def logo_png():
    return redirect(url_for('static', filename='logo.png'))

if __name__ == '__main__':
    # Create 'static' directory if it doesn't exist for map.html
    os.makedirs(os.path.join(app.root_path, 'static'), exist_ok=True)
    # Create 'templates' directory if it doesn't exist
    os.makedirs(os.path.join(app.root_path, 'templates'), exist_ok=True)
    # Ensure reports file exists
    if not os.path.exists(REPORTS_FILE):
        with open(REPORTS_FILE, 'w', encoding='utf-8') as f:
            json.dump([], f)
    ensure_image_folders()
    print("Starting Flask app on http://127.0.0.1:5000/")
    print("DEBUG: Waiting for POST /submit_report requests. If you do not see any after submitting a report, check your HTML/JS form submission.")
    app.run(debug=True) # debug=True allows auto-reloading and detailed error messages